#!/usr/bin/env python3
"""
Extractor de Facturas - Informacion Exogena 2025
Formato 1001: Pagos o abonos en cuenta
Año gravable 2025 - Persona natural, declarante de renta

Uso:
    python extractor_facturas.py                        # carpeta ./facturas por defecto
    python extractor_facturas.py "C:/mis facturas"      # carpeta especifica
"""

import os
import re
import sys
import logging
from pathlib import Path
from datetime import datetime
import xml.etree.ElementTree as ET
from collections import defaultdict
from typing import Optional, Dict, List

# Forzar UTF-8 en la consola de Windows
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() != "utf-8":
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

# ── Dependencias requeridas ───────────────────────────────────────────────────
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: Instalar dependencias basicas:\n  pip install pandas openpyxl")

# ── Dependencias opcionales ───────────────────────────────────────────────────
MISSING = []

try:
    import pdfplumber
except ImportError:
    pdfplumber = None
    MISSING.append("pdfplumber  (para leer PDFs)")

try:
    from PIL import Image
    import pytesseract
except ImportError:
    Image = None
    pytesseract = None
    MISSING.append("Pillow pytesseract  (para leer imagenes JPG/PNG)")

# Rutas comunes de Tesseract en Windows
if pytesseract:
    for _p in [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        r"C:\Users\User\AppData\Local\Programs\Tesseract-OCR\tesseract.exe",
    ]:
        if os.path.exists(_p):
            pytesseract.pytesseract.tesseract_cmd = _p
            break

# ── Namespaces DIAN (factura electronica UBL 2.1) ────────────────────────────
NS = {
    "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
    "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
    "ext": "urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2",
}

# ── Patrones regex ────────────────────────────────────────────────────────────
NIT_RE = re.compile(
    r"(?:N\.?I\.?T\.?)\s*[:\-\s.]*(\d{6,11})\s*[-–]?\s*(\d?)", re.IGNORECASE
)
CC_RE = re.compile(r"(?:C\.?C\.?|Cedula|Cedula de Ciudadania)\s*[:\s]*(\d{6,12})", re.IGNORECASE)

DATE_ISO = re.compile(r"(\d{4})-(\d{2})-(\d{2})")
DATE_COL = re.compile(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})")
DATE_ESP = re.compile(
    r"(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|"
    r"septiembre|octubre|noviembre|diciembre)\s+de\s+(\d{4})",
    re.IGNORECASE,
)
MESES = {
    "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
    "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
    "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12",
}

TOTAL_RE = re.compile(
    r"(?:TOTAL\s+A\s+PAGAR|VALOR\s+TOTAL|TOTAL\s+FACTURA|TOTAL\s+NETO|"
    r"VALOR\s+A\s+PAGAR|GRAND\s+TOTAL|TOTAL\s*:)\s*\$?\s*([\d.,]+)",
    re.IGNORECASE,
)
BASE_RE = re.compile(
    r"(?:SUBTOTAL|BASE\s+GRAVABLE|BASE\s+IVA|VALOR\s+ANTES\s+(?:DE\s+)?IVA|"
    r"VALOR\s+BASE|BASE\s+IMPONIBLE)\s*:?\s*\$?\s*([\d.,]+)",
    re.IGNORECASE,
)
IVA_RE = re.compile(
    r"(?:I\.?V\.?A\.?|IMPUESTO\s+AL\s+VALOR\s+AGREGADO)\s*(?:\(?\d+\s*%\)?)?\s*[:\$]\s*([\d.,]+)",
    re.IGNORECASE,
)
NOMBRE_RE = re.compile(
    r"(?:RAZON\s+SOCIAL|RAZ[OÓ]N\s+SOCIAL|EXPEDIDO\s+POR|EMISOR|PROVEEDOR|EMPRESA)\s*[:\-]?\s*([^\n\r]{3,80})",
    re.IGNORECASE,
)
INVOICE_RE = re.compile(
    r"(?:FACTURA\s+(?:DE\s+VENTA\s+)?(?:N[oO°]?\.?|N[uU]?M\.?)|FV|FA|FE)\s*[:\-\s]?\s*([A-Z0-9\-]{3,20})",
    re.IGNORECASE,
)

CONCEPTOS = {
    "5002": ["honorario", "honorarios", "asesoria legal", "asesoría legal", "contable"],
    "5003": ["servicio", "servicios", "soporte", "consultoria", "consultoría",
             "asesoria", "asesoría", "mantenimiento", "capacitacion", "diseño web",
             "desarrollo", "publicidad", "marketing"],
    "5004": ["arrendamiento", "arriendo", "alquiler", "renta local", "canon"],
    "5005": ["intereses", "rendimientos financieros", "credito", "prestamo"],
}

IMG_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp"}


# ─────────────────────────────────────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────────────────────────────────────

def clean_nit(raw: str) -> str:
    return re.sub(r"[^\d]", "", raw or "")


def calc_dv(nit: str) -> str:
    """Calcula digito de verificacion de NIT colombiano."""
    nit = clean_nit(nit)
    if not nit:
        return ""
    factores = [71, 67, 59, 53, 47, 43, 41, 37, 29, 23, 19, 17, 13, 7, 3]
    nit_pad = nit.zfill(15)
    total = sum(int(d) * f for d, f in zip(nit_pad, factores))
    r = total % 11
    return str(0 if r in (0, 1) else 11 - r)


def parse_amount(text: str) -> float:
    """Convierte texto de monto (formato colombiano o americano) a float."""
    if not text:
        return 0.0
    t = re.sub(r"[^\d.,]", "", str(text))
    if not t:
        return 0.0
    dots, commas = t.count("."), t.count(",")
    if dots == 0 and commas == 0:
        return float(t)
    if dots > 0 and commas > 0:
        # El ultimo separador es el decimal
        if t.rfind(",") > t.rfind("."):        # 1.234.567,89 → EU/colombiano
            t = t.replace(".", "").replace(",", ".")
        else:                                   # 1,234,567.89 → americano
            t = t.replace(",", "")
    elif commas > 0 and dots == 0:
        parts = t.split(",")
        if len(parts) == 2 and len(parts[-1]) == 2:
            t = t.replace(",", ".")             # 1234567,89 → decimal
        else:
            t = t.replace(",", "")             # 1,234,567 → miles
    else:
        parts = t.split(".")
        if len(parts) == 2 and len(parts[-1]) == 2:
            pass                               # 1234567.89 → decimal, OK
        else:
            t = t.replace(".", "")             # 1.234.567 → miles
    try:
        return float(t)
    except ValueError:
        return 0.0


def normalizar_fecha(raw: str) -> str:
    """Normaliza cualquier formato de fecha a DD/MM/YYYY."""
    if not raw:
        return ""
    m = DATE_ESP.search(raw)
    if m:
        return f"{m.group(1).zfill(2)}/{MESES[m.group(2).lower()]}/{m.group(3)}"
    m = DATE_ISO.search(raw)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    m = DATE_COL.search(raw)
    if m:
        d, mo, y = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        return f"{d}/{mo}/{y}"
    return raw.strip()


def detect_concepto(text: str) -> str:
    """Infiere el concepto DIAN 1001 segun palabras clave del texto."""
    t = text.lower()
    for code, keywords in CONCEPTOS.items():
        if any(k in t for k in keywords):
            return code
    return "5001"  # Compras por defecto


def empty_record(path: Path) -> Dict:
    return {
        "archivo": path.name,
        "numero_factura": "",
        "tipo_doc": "31",
        "nit": "",
        "dv": "",
        "proveedor": "",
        "fecha": "",
        "base": 0.0,
        "iva": 0.0,
        "total": 0.0,
        "concepto": "5001",
        "fuente": path.suffix.upper().lstrip("."),
        "confianza": "Baja",
        "error": "",
    }


def score_confidence(rec: Dict) -> str:
    filled = sum(1 for f in ["nit", "proveedor", "fecha", "total"] if rec.get(f))
    if filled == 4:
        return "Alta"
    if filled >= 2:
        return "Media"
    return "Baja"


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACCION DE CAMPOS DESDE TEXTO PLANO (PDF e imagen)
# ─────────────────────────────────────────────────────────────────────────────

def extract_fields_from_text(text: str, rec: Dict) -> Dict:
    """Aplica regex al texto extraido para rellenar campos."""

    # Numero de factura
    if not rec.get("numero_factura"):
        m = INVOICE_RE.search(text)
        if m:
            rec["numero_factura"] = m.group(1).strip()

    # NIT proveedor
    if not rec.get("nit"):
        m = NIT_RE.search(text)
        if m:
            rec["nit"] = m.group(1)
            rec["dv"] = m.group(2) or calc_dv(m.group(1))
            rec["tipo_doc"] = "31"
        else:
            m = CC_RE.search(text)
            if m:
                rec["nit"] = m.group(1)
                rec["tipo_doc"] = "13"

    # Nombre proveedor
    if not rec.get("proveedor"):
        m = NOMBRE_RE.search(text)
        if m:
            rec["proveedor"] = m.group(1).strip()[:100]

    # Fecha
    if not rec.get("fecha"):
        for pat in [DATE_ESP, DATE_ISO, DATE_COL]:
            m = pat.search(text)
            if m:
                rec["fecha"] = normalizar_fecha(m.group(0))
                break

    # Montos
    if not rec.get("total"):
        m = TOTAL_RE.search(text)
        if m:
            rec["total"] = parse_amount(m.group(1))

    if not rec.get("base"):
        m = BASE_RE.search(text)
        if m:
            rec["base"] = parse_amount(m.group(1))

    if not rec.get("iva"):
        m = IVA_RE.search(text)
        if m:
            rec["iva"] = parse_amount(m.group(1))

    # Inferencias
    if rec.get("total") and rec.get("iva") and not rec.get("base"):
        rec["base"] = round(rec["total"] - rec["iva"], 0)
    elif rec.get("total") and not rec.get("base") and not rec.get("iva"):
        rec["base"] = rec["total"]

    # Concepto
    rec["concepto"] = detect_concepto(text)

    return rec


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTOR XML (Factura electronica DIAN UBL 2.1)
# ─────────────────────────────────────────────────────────────────────────────

def extract_from_xml(path: Path) -> Dict:
    rec = empty_record(path)
    rec["fuente"] = "XML"

    try:
        tree = ET.parse(path)
        root = tree.getroot()

        def find(xpath: str) -> Optional[str]:
            el = root.find(xpath, NS)
            return el.text.strip() if el is not None and el.text else None

        def find_attr(xpath: str, attr: str) -> str:
            el = root.find(xpath, NS)
            return el.get(attr, "").strip() if el is not None else ""

        rec["numero_factura"] = find(".//cbc:ID") or ""
        rec["fecha"] = normalizar_fecha(find(".//cbc:IssueDate") or "")

        # Proveedor (AccountingSupplierParty = el que emite = proveedor)
        nit_raw = find(".//cac:AccountingSupplierParty//cbc:CompanyID") or ""
        rec["nit"] = clean_nit(nit_raw)
        rec["dv"] = find_attr(".//cac:AccountingSupplierParty//cbc:CompanyID", "schemeID")
        if not rec["dv"] and rec["nit"]:
            rec["dv"] = calc_dv(rec["nit"])

        nombre = (
            find(".//cac:AccountingSupplierParty//cbc:RegistrationName")
            or find(".//cac:AccountingSupplierParty//cbc:Name")
            or ""
        )
        rec["proveedor"] = nombre[:100]

        scheme = find_attr(".//cac:AccountingSupplierParty//cbc:CompanyID", "schemeName")
        rec["tipo_doc"] = "31" if "NIT" in scheme.upper() else "13"

        # Montos
        base_raw = find(".//cac:LegalMonetaryTotal/cbc:LineExtensionAmount") or "0"
        iva_raw = find(".//cac:TaxTotal/cbc:TaxAmount") or "0"
        total_raw = find(".//cac:LegalMonetaryTotal/cbc:PayableAmount") or "0"

        rec["base"] = round(float(base_raw), 0)
        rec["iva"] = round(float(iva_raw), 0)
        rec["total"] = round(float(total_raw), 0)

        # Concepto desde descripcion de lineas
        desc = find(".//cac:InvoiceLine//cbc:Description") or find(".//cac:Item/cbc:Name") or ""
        rec["concepto"] = detect_concepto(desc)

        rec["confianza"] = score_confidence(rec)

    except Exception as e:
        log.warning(f"XML {path.name}: {e}")
        rec["confianza"] = "Error"
        rec["error"] = str(e)

    return rec


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTOR PDF
# ─────────────────────────────────────────────────────────────────────────────

def extract_from_pdf(path: Path) -> Dict:
    rec = empty_record(path)
    rec["fuente"] = "PDF"

    if pdfplumber is None:
        rec["confianza"] = "Error"
        rec["error"] = "pdfplumber no instalado. Ejecutar: pip install pdfplumber"
        return rec

    try:
        pages_text = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    pages_text.append(t)

        full_text = "\n".join(pages_text)

        if not full_text.strip():
            # PDF basado en imagen: intentar OCR si esta disponible
            if pytesseract and Image:
                log.info(f"  PDF sin texto → intentando OCR: {path.name}")
                rec["fuente"] = "PDF+OCR"
                with pdfplumber.open(path) as pdf:
                    imgs = []
                    for page in pdf.pages:
                        img = page.to_image(resolution=200).original
                        imgs.append(img)
                if imgs:
                    full_text = "\n".join(
                        pytesseract.image_to_string(img, lang="spa", config="--psm 6")
                        for img in imgs
                    )
            if not full_text.strip():
                rec["confianza"] = "Error"
                rec["error"] = "PDF sin texto extraible (imagen sin OCR instalado)"
                return rec

        rec = extract_fields_from_text(full_text, rec)
        rec["confianza"] = score_confidence(rec)

    except Exception as e:
        log.warning(f"PDF {path.name}: {e}")
        rec["confianza"] = "Error"
        rec["error"] = str(e)

    return rec


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTOR IMAGEN (OCR)
# ─────────────────────────────────────────────────────────────────────────────

def extract_from_image(path: Path) -> Dict:
    rec = empty_record(path)
    rec["fuente"] = "IMG"

    if pytesseract is None or Image is None:
        rec["confianza"] = "Error"
        rec["error"] = "pytesseract/Pillow no instalado. Ejecutar: pip install Pillow pytesseract"
        return rec

    try:
        img = Image.open(path).convert("L")  # Escala de grises
        w, h = img.size
        # Aumentar resolucion si la imagen es pequena
        if max(w, h) < 1800:
            scale = 1800 / max(w, h)
            img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)

        text = pytesseract.image_to_string(img, lang="spa", config="--psm 6")

        if not text.strip():
            rec["confianza"] = "Error"
            rec["error"] = "OCR no extrajo texto. Verificar calidad de imagen."
            return rec

        rec = extract_fields_from_text(text, rec)
        rec["confianza"] = score_confidence(rec)

    except Exception as e:
        log.warning(f"IMG {path.name}: {e}")
        rec["confianza"] = "Error"
        rec["error"] = str(e)

    return rec


# ─────────────────────────────────────────────────────────────────────────────
# PROCESADOR DE CARPETA
# ─────────────────────────────────────────────────────────────────────────────

def process_folder(folder: Path, year_filter: int = 2025) -> List[Dict]:
    """Procesa todos los archivos de facturas en la carpeta."""
    # Preferir XML sobre PDF del mismo nombre base
    all_files = list(folder.iterdir())
    xml_stems = {f.stem.lower() for f in all_files if f.suffix.lower() == ".xml"}

    records = []
    skipped = []

    for f in sorted(all_files):
        ext = f.suffix.lower()
        if not f.is_file():
            continue

        if ext == ".xml":
            rec = extract_from_xml(f)
        elif ext == ".pdf":
            # Saltar PDF si existe XML con el mismo nombre
            if f.stem.lower() in xml_stems:
                log.info(f"  Saltando {f.name} (usando XML del mismo nombre)")
                continue
            rec = extract_from_pdf(f)
        elif ext in IMG_EXTS:
            rec = extract_from_image(f)
        else:
            skipped.append(f.name)
            continue

        # Filtro de año
        if rec.get("fecha") and year_filter:
            parts = rec["fecha"].split("/")
            if len(parts) == 3:
                try:
                    if int(parts[2]) != year_filter:
                        rec["error"] = f"Fecha {rec['fecha']} fuera del año {year_filter}"
                        rec["confianza"] = "Revisar"
                except ValueError:
                    pass

        records.append(rec)
        icon = "OK" if rec["confianza"] in ("Alta", "Media") else "!!"
        prov = (rec.get("proveedor") or "?")[:35]
        total = rec.get("total") or 0
        log.info(f"  {icon} {f.name:<35} | {prov:<35} | ${total:>14,.0f}")

    if skipped:
        log.info(f"\n  Ignorados (formato no soportado): {', '.join(skipped[:8])}")

    return records


# ─────────────────────────────────────────────────────────────────────────────
# GENERACION DEL EXCEL
# ─────────────────────────────────────────────────────────────────────────────

C_AZUL    = "1F3864"
C_AZUL2   = "2F5496"
C_NARANJA = "FFE0B2"
C_ROJO    = "FFCDD2"
C_VERDE   = "E8F5E9"
C_GRIS    = "F5F5F5"
C_AMARILLO= "FFF9C4"


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(cell, bg: str = C_AZUL):
    cell.fill = _fill(bg)
    cell.font = Font(bold=True, color="FFFFFF", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()


def generate_excel(records: List[Dict], output: Path):
    wb = Workbook()

    # ── Hoja 1: Formato 1001 ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Formato 1001"
    ws.sheet_view.showGridLines = False

    # Titulo
    ws.merge_cells("A1:R1")
    ws["A1"] = "INFORMACION EXOGENA 2025 — FORMATO 1001: PAGOS O ABONOS EN CUENTA"
    ws["A1"].font = Font(bold=True, size=13, color=C_AZUL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:R2")
    ws["A2"] = (
        f"Persona Natural | Declarante de renta | Solo retenciones en cero "
        f"(no es agente retenedor) | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws["A2"].font = Font(italic=True, size=8, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 14

    # Encabezados
    headers = [
        ("Archivo",                             20),
        ("N° Factura",                          14),
        ("Tipo Doc\n31=NIT / 13=CC",            10),
        ("NIT o Cedula\nProveedor",             16),
        ("DV",                                   5),
        ("Razon Social / Nombre\nProveedor",    38),
        ("Fecha\nFactura",                      11),
        ("Concepto\n5001 Compras\n5002 Honor.\n5003 Servicios\n5004 Arrend.", 12),
        ("Valor Base\n(sin IVA)",               16),
        ("IVA",                                 14),
        ("Valor Total\nFactura",                14),
        ("Ret. Fuente\n(0)",                     9),
        ("Ret. IVA\n(0)",                        9),
        ("Ret. Timbre\n(0)",                     9),
        ("Pais\n169=CO",                         8),
        ("Fuente",                               8),
        ("Confianza",                            10),
        ("Observaciones",                        30),
    ]

    ROW_H = 3
    for col, (hdr, width) in enumerate(headers, 1):
        cell = ws.cell(row=ROW_H, column=col, value=hdr)
        _header_cell(cell)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[ROW_H].height = 58

    # Filas de datos
    fill_alert   = _fill(C_NARANJA)
    fill_error   = _fill(C_ROJO)
    fill_ok      = _fill("FFFFFF")
    fill_revisar = _fill(C_AMARILLO)

    for row_idx, rec in enumerate(records, ROW_H + 1):
        conf = rec.get("confianza", "Baja")
        bg = (
            fill_error   if conf == "Error"
            else fill_alert   if conf == "Baja"
            else fill_revisar if conf == "Revisar"
            else fill_ok
        )
        obs = rec.get("error") or (
            "Revisar concepto manualmente" if rec.get("concepto") == "5001" else ""
        )

        row_vals = [
            rec["archivo"],
            rec["numero_factura"],
            rec["tipo_doc"],
            rec["nit"],
            rec["dv"],
            rec["proveedor"],
            rec["fecha"],
            rec["concepto"],
            rec["base"] or None,
            rec["iva"] or None,
            rec["total"] or None,
            0,          # Ret. fuente
            0,          # Ret. IVA
            0,          # Ret. timbre
            "169",      # Colombia
            rec["fuente"],
            conf,
            obs,
        ]

        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = bg
            cell.border = _border()
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if col in (9, 10, 11):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if col in (3, 4, 5, 12, 13, 14, 15):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fila de totales
    total_row = len(records) + ROW_H + 1
    ws.cell(row=total_row, column=8, value="TOTALES").font = Font(bold=True)
    for col, field in [(9, "base"), (10, "iva"), (11, "total")]:
        v = sum((r.get(field) or 0) for r in records if r.get("confianza") not in ("Error",))
        cell = ws.cell(row=total_row, column=col, value=v)
        cell.font = Font(bold=True)
        cell.number_format = "#,##0"
        cell.fill = _fill("DDEEFF")
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _border()

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:R{len(records) + ROW_H}"

    # ── Hoja 2: Resumen por proveedor ─────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen por Proveedor")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:G1")
    ws2["A1"] = "RESUMEN CONSOLIDADO POR PROVEEDOR — Año gravable 2025"
    ws2["A1"].font = Font(bold=True, size=12, color=C_AZUL)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24

    h2_cols = [
        ("NIT / Cedula", 16), ("DV", 5), ("Razon Social / Nombre", 40),
        ("Tipo Doc", 10), ("N° Facturas", 12),
        ("Total Base", 16), ("Total IVA", 14), ("Total Pagado", 16),
    ]
    for col, (h, w) in enumerate(h2_cols, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        _header_cell(cell, C_AZUL2)
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[2].height = 22

    grupos: Dict[str, Dict] = defaultdict(
        lambda: {"proveedor": "", "dv": "", "tipo_doc": "31", "count": 0,
                 "base": 0.0, "iva": 0.0, "total": 0.0}
    )
    for r in records:
        if r.get("confianza") in ("Error",) or not r.get("nit"):
            continue
        k = r["nit"]
        g = grupos[k]
        g["proveedor"] = g["proveedor"] or r["proveedor"]
        g["dv"]        = g["dv"]        or r["dv"]
        g["tipo_doc"]  = r["tipo_doc"]
        g["count"]    += 1
        g["base"]     += r.get("base")  or 0
        g["iva"]      += r.get("iva")   or 0
        g["total"]    += r.get("total") or 0

    for row_idx, (nit, d) in enumerate(sorted(grupos.items()), 3):
        row_data = [nit, d["dv"], d["proveedor"], d["tipo_doc"], d["count"],
                    d["base"], d["iva"], d["total"]]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.border = _border()
            cell.alignment = Alignment(vertical="center")
            if col in (6, 7, 8):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if col in (1, 2, 4, 5):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Hoja 3: Pendientes de revision ───────────────────────────────────────
    pendientes = [
        r for r in records
        if r.get("confianza") in ("Baja", "Error", "Revisar") or not r.get("nit")
    ]
    if pendientes:
        ws3 = wb.create_sheet("Pendientes Revision")
        ws3.sheet_view.showGridLines = False
        ws3.merge_cells("A1:F1")
        ws3["A1"] = f"FACTURAS QUE REQUIEREN REVISION MANUAL ({len(pendientes)})"
        ws3["A1"].font = Font(bold=True, size=12, color="C62828")
        ws3["A1"].alignment = Alignment(horizontal="center")
        ws3.row_dimensions[1].height = 24

        h3 = [("Archivo", 32), ("Confianza", 12), ("NIT", 16), ("Proveedor", 35),
              ("Total", 14), ("Motivo / Error", 50)]
        for col, (h, w) in enumerate(h3, 1):
            cell = ws3.cell(row=2, column=col, value=h)
            _header_cell(cell, "C62828")
            ws3.column_dimensions[get_column_letter(col)].width = w

        for row_idx, r in enumerate(pendientes, 3):
            motivo = r.get("error") or "Datos incompletos — completar manualmente"
            row_data = [r["archivo"], r["confianza"], r["nit"], r["proveedor"],
                        r.get("total") or 0, motivo]
            for col, val in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col, value=val)
                cell.border = _border()
                cell.fill = _fill("FFF3E0")
                if col == 5:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")

    # ── Hoja 4: Instrucciones ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Instrucciones")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 80

    instrucciones = [
        ("INSTRUCCIONES PARA USAR ESTE ARCHIVO", True, C_AZUL, "FFFFFF", 14),
        ("", False, None, None, 12),
        ("1. COLUMNA 'CONCEPTO' — Codigos DIAN Formato 1001:", True, C_AZUL2, "FFFFFF", 10),
        ("   5001 = Compras de bienes", False, None, "000000", 10),
        ("   5002 = Honorarios", False, None, "000000", 10),
        ("   5003 = Servicios", False, None, "000000", 10),
        ("   5004 = Arrendamientos", False, None, "000000", 10),
        ("   5005 = Intereses y rendimientos financieros", False, None, "000000", 10),
        ("   5006 = Otros pagos", False, None, "000000", 10),
        ("", False, None, None, 10),
        ("2. TIPO DE DOCUMENTO del proveedor:", True, C_AZUL2, "FFFFFF", 10),
        ("   31 = NIT (personas juridicas / empresas)", False, None, "000000", 10),
        ("   13 = Cedula de ciudadania (persona natural colombiana)", False, None, "000000", 10),
        ("   22 = Cedula de extranjeria", False, None, "000000", 10),
        ("   41 = Pasaporte", False, None, "000000", 10),
        ("", False, None, None, 10),
        ("3. RETENCIONES en cero (columnas J, K, L):", True, C_AZUL2, "FFFFFF", 10),
        ("   Su cliente es persona natural NO agente retenedor.", False, None, "000000", 10),
        ("   Por eso todas las retenciones van en CERO.", False, None, "000000", 10),
        ("", False, None, None, 10),
        ("4. TOPES para reportar (verificar con la DIAN para 2025):", True, C_AZUL2, "FFFFFF", 10),
        ("   Solo reportar proveedores que superen el tope minimo.", False, None, "000000", 10),
        ("   Para 2024 era $100.000 por proveedor. Confirmar tope 2025.", False, None, "000000", 10),
        ("", False, None, None, 10),
        ("5. FILAS NARANJAS = baja confianza en extraccion. Completar manualmente.", True, "E65100", "FFFFFF", 10),
        ("6. FILAS ROJAS = error de procesamiento. Ingresar datos a mano.", True, "C62828", "FFFFFF", 10),
        ("", False, None, None, 10),
        ("7. Pais Colombia = codigo 169", False, None, "000000", 10),
        ("8. Este archivo es un borrador de trabajo. Siempre valide con su contador.", False, None, "666666", 9),
    ]

    for row_idx, (text, bold, bg, fg, size) in enumerate(instrucciones, 1):
        cell = ws4.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold, size=size, color=fg or "000000")
        cell.alignment = Alignment(vertical="center")
        ws4.row_dimensions[row_idx].height = 18
        if bg:
            cell.fill = _fill(bg)

    wb.save(output)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def iter_folder(folder: Path, year_filter: int = 2025):
    """Generador que procesa archivos uno por uno. Yield de (record, idx, total)."""
    if not folder.is_dir():
        return
    all_files = [f for f in sorted(folder.iterdir()) if f.is_file()]
    xml_stems = {f.stem.lower() for f in all_files if f.suffix.lower() == ".xml"}

    processable = []
    for f in all_files:
        ext = f.suffix.lower()
        if ext == ".xml":
            processable.append(f)
        elif ext == ".pdf" and f.stem.lower() not in xml_stems:
            processable.append(f)
        elif ext in IMG_EXTS:
            processable.append(f)

    total = len(processable)
    for idx, f in enumerate(processable, 1):
        ext = f.suffix.lower()
        if ext == ".xml":
            rec = extract_from_xml(f)
        elif ext == ".pdf":
            rec = extract_from_pdf(f)
        else:
            rec = extract_from_image(f)

        if rec.get("fecha") and year_filter:
            parts = rec["fecha"].split("/")
            if len(parts) == 3:
                try:
                    if int(parts[2]) != year_filter:
                        rec["error"] = f"Fecha {rec['fecha']} no corresponde al ano {year_filter}"
                        rec["confianza"] = "Revisar"
                except ValueError:
                    pass

        yield rec, idx, total


def main():
    folder = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("facturas")

    print("\n" + "="*65)
    print("  EXTRACTOR DE FACTURAS - INFORMACION EXOGENA 2025 (Formato 1001)")
    print("="*65 + "\n")

    if MISSING:
        print("  Dependencias opcionales no instaladas:")
        for dep in MISSING:
            print(f"    pip install {dep}")
        print()

    if not folder.exists():
        print(f"  ERROR: La carpeta '{folder}' no existe.\n")
        print(f"  Opciones:")
        print(f"    1. Crea la carpeta 'facturas' en esta misma ubicacion")
        print(f"       y coloca ahi todas las facturas (PDF, XML, JPG, PNG).")
        print(f"    2. O pasa la ruta como argumento:")
        print(f'       python extractor_facturas.py "C:/ruta/a/mis/facturas"\n')
        sys.exit(1)

    print(f"  Carpeta: {folder.resolve()}\n")
    print(f"  {'Archivo':<35} | {'Proveedor':<35} | {'Total':>14}")
    print("  " + "-"*35 + "-+-" + "-"*35 + "-+-" + "-"*14)

    records = process_folder(folder, year_filter=2025)

    if not records:
        print("\n  No se encontraron archivos para procesar.")
        print("  Formatos soportados: PDF, XML, JPG, JPEG, PNG, BMP, TIFF\n")
        sys.exit(0)

    total_ok      = sum(1 for r in records if r["confianza"] in ("Alta", "Media"))
    total_revisar = sum(1 for r in records if r["confianza"] in ("Baja", "Revisar"))
    total_error   = sum(1 for r in records if r["confianza"] == "Error")
    gran_total    = sum((r.get("total") or 0) for r in records if r["confianza"] != "Error")
    gran_iva      = sum((r.get("iva") or 0) for r in records if r["confianza"] != "Error")
    gran_base     = sum((r.get("base") or 0) for r in records if r["confianza"] != "Error")

    print("\n" + "-"*65)
    print(f"  Archivos procesados : {len(records)}")
    print(f"  Extraidos bien       : {total_ok}")
    print(f"  Requieren revision   : {total_revisar}")
    print(f"  Con errores          : {total_error}")
    print("  " + "-"*43)
    print(f"  Total base (sin IVA) : ${gran_base:>18,.0f}")
    print(f"  Total IVA            : ${gran_iva:>18,.0f}")
    print(f"  TOTAL FACTURAS       : ${gran_total:>18,.0f}")
    print("-"*65 + "\n")

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    output = folder.parent / f"exogena_2025_F1001_{ts}.xlsx"
    generate_excel(records, output)

    print(f"  Archivo generado: {output.resolve()}\n")
    print("  RECUERDE:")
    print("  - Revise el concepto de cada fila (5001/5002/5003/5004)")
    print("  - Corrija filas en naranja/rojo en la hoja 'Pendientes Revision'")
    print("  - Verifique NITs y DV antes de presentar a la DIAN")
    print("  - Consulte los topes de reporte vigentes para año gravable 2025")
    print("="*65 + "\n")


if __name__ == "__main__":
    main()
